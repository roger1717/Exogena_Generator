# app/services/csv_processor.py

import pandas as pd
import re
from typing import Dict, List, Any, Tuple, Optional
from sqlalchemy.orm import Session
from datetime import datetime
from pathlib import Path

from app.models.mapping_rule import MappingRule
from app.schemas.exogena import ProcessingResult

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class CSVProcessor:
    """
    Servicio para procesar archivos CSV con detección automática de columnas.
    """

    # Mapeo predefinido para columnas comunes (puedes expandirlo)
    COLUMN_MAPPING_TEMPLATES = {
        'auxiliar_contable': {
            'nit': 'NIT_Tercero',
            'nombre': 'Razon_Social',
            'valor': '__calcular_debito_credito__',
            'debito': 'Debito',
            'credito': 'Credito',
            'puc': 'Cuenta',
            'fecha': 'Fecha',
            'concepto': 'Concepto'
        }
    }

    def __init__(self, db: Session, column_mapping: Optional[Dict[str, str]] = None):
        self.db = db
        self.column_mapping = column_mapping

    def _detect_column_mapping(self, df: pd.DataFrame) -> Dict[str, str]:
        """Detecta automáticamente el mapeo de columnas."""
        columns = df.columns.tolist()
        mapping = {}

        # Intentar con plantillas predefinidas
        for template_name, template in self.COLUMN_MAPPING_TEMPLATES.items():
            match_count = 0
            for field, value in template.items():
                if field == 'valor' and value == '__calcular_debito_credito__':
                    # Verificar si existen débito y crédito
                    if 'debito' in template and 'credito' in template:
                        for col in columns:
                            col_lower = col.lower().strip()
                            if 'debito' in col_lower or 'debe' in col_lower or 'cargo' in col_lower:
                                mapping['debito'] = col
                            if 'credito' in col_lower or 'haber' in col_lower or 'abono' in col_lower:
                                mapping['credito'] = col
                        if 'debito' in mapping and 'credito' in mapping:
                            mapping['valor'] = '__calcular_debito_credito__'
                            match_count += 2
                else:
                    # Buscar coincidencia exacta o parcial
                    for col in columns:
                        col_lower = col.lower().strip()
                        if value.lower() in col_lower or col_lower == value.lower():
                            mapping[field] = col
                            match_count += 1
                            break

            # Si encontramos al menos 3 campos requeridos, usamos esta plantilla
            required = ['nit', 'nombre', 'valor', 'puc']
            found = [f for f in required if f in mapping]
            if len(found) >= 3:
                return mapping

        # Si no funcionó, usar búsqueda por patrones genéricos
        patterns = {
            'nit': [r'nit', r'n\.i\.t', r'identificacion', r'cedula', r'ruc', r'documento', r'tercero'],
            'nombre': [r'nombre', r'razon', r'social', r'cliente', r'proveedor', r'beneficiario'],
            'valor': [r'valor', r'monto', r'importe', r'cantidad', r'debito', r'credito'],
            'puc': [r'cuenta', r'puc', r'codigo', r'cod', r'cta'],
            'fecha': [r'fecha', r'date', r'dia'],
            'concepto': [r'concepto', r'descripcion', r'detalle', r'glosa']
        }

        for field, pattern_list in patterns.items():
            for col in columns:
                col_lower = col.lower().strip()
                for pattern in pattern_list:
                    if re.search(pattern, col_lower, re.IGNORECASE):
                        mapping[field] = col
                        break
                if field in mapping:
                    break

        # Detectar débito y crédito si no se detectó valor
        if 'valor' not in mapping:
            for col in columns:
                col_lower = col.lower().strip()
                if 'debito' in col_lower or 'debe' in col_lower or 'cargo' in col_lower:
                    mapping['debito'] = col
                if 'credito' in col_lower or 'haber' in col_lower or 'abono' in col_lower:
                    mapping['credito'] = col
            if 'debito' in mapping and 'credito' in mapping:
                mapping['valor'] = '__calcular_debito_credito__'

        # Si aún no hay valor, tomar la primera columna numérica
        if 'valor' not in mapping:
            numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
            if numeric_cols:
                mapping['valor'] = numeric_cols[0]

        # Validar campos requeridos
        required_fields = ['nit', 'nombre', 'valor', 'puc']
        missing = [f for f in required_fields if f not in mapping or mapping[f] is None]
        if missing:
            raise ValueError(f"No se detectaron las columnas: {missing}. Columnas disponibles: {columns}")

        return mapping

    def _load_csv(self, file_path: str) -> pd.DataFrame:
        """Carga CSV probando diferentes codificaciones."""
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                df.columns = df.columns.str.strip()
                return df
            except:
                continue
        raise ValueError(f"No se pudo leer el archivo {file_path}")

    def _normalize_dataframe(self, df: pd.DataFrame, column_map: Dict[str, str]) -> pd.DataFrame:
        """Normaliza el DataFrame usando el mapeo."""
        df_norm = pd.DataFrame()
        if column_map.get('nit'):
            df_norm['nit_tercero'] = df[column_map['nit']].astype(str).str.strip()
        if column_map.get('nombre'):
            df_norm['nombre_tercero'] = df[column_map['nombre']].astype(str).str.strip()
        if column_map.get('puc'):
            df_norm['codigo_puc'] = df[column_map['puc']].astype(str).str.strip()

        # Valor
        if column_map.get('valor') == '__calcular_debito_credito__':
            debito_col = column_map.get('debito')
            credito_col = column_map.get('credito')
            if debito_col and credito_col:
                df_norm['valor'] = pd.to_numeric(df[debito_col], errors='coerce').fillna(0) + \
                                   pd.to_numeric(df[credito_col], errors='coerce').fillna(0)
            else:
                df_norm['valor'] = 0
        elif column_map.get('valor'):
            df_norm['valor'] = pd.to_numeric(df[column_map['valor']], errors='coerce').fillna(0)
        else:
            df_norm['valor'] = 0

        # Fecha y concepto (opcionales)
        if column_map.get('fecha'):
            try:
                df_norm['fecha'] = pd.to_datetime(df[column_map['fecha']])
            except:
                df_norm['fecha'] = None
        if column_map.get('concepto'):
            df_norm['concepto'] = df[column_map['concepto']].astype(str)

        # Limpiar
        df_norm = df_norm.dropna(subset=['nit_tercero', 'codigo_puc'])
        df_norm = df_norm[df_norm['nit_tercero'] != '']
        df_norm = df_norm[df_norm['nit_tercero'] != 'nan']
        df_norm = df_norm[df_norm['codigo_puc'] != '']
        df_norm = df_norm[df_norm['codigo_puc'] != 'nan']
        df_norm = df_norm[df_norm['valor'] != 0]
        return df_norm

    def _process_rows(self, df: pd.DataFrame, rule_map: Dict[str, Any]) -> Tuple[List[Dict], List[Dict]]:
        """Procesa filas y aplica reglas de mapeo."""
        processed = []
        errors = []
        for idx, row in df.iterrows():
            row_num = idx + 2
            nit = str(row.get('nit_tercero', '')).strip()
            nombre = str(row.get('nombre_tercero', '')).strip()
            valor = float(row.get('valor', 0))
            puc = str(row.get('codigo_puc', '')).strip()

            if not nit or nit == 'nan':
                errors.append({'row': row_num, 'error': 'NIT vacío'})
                continue
            if valor == 0:
                errors.append({'row': row_num, 'error': 'Valor cero'})
                continue
            if not puc or puc == 'nan':
                errors.append({'row': row_num, 'error': 'PUC vacío', 'data': {'nit': nit, 'nombre': nombre, 'valor': valor}})
                continue

            rule = rule_map.get(puc)
            if not rule:
                errors.append({'row': row_num, 'error': f'PUC {puc} sin regla', 'data': {'nit': nit, 'nombre': nombre, 'valor': valor}})
                continue

            processed.append({
                'nit_tercero': nit,
                'nombre_tercero': nombre,
                'valor': valor,
                'codigo_puc': puc,
                'concepto_asignado': rule.exogena_concept,
                'formato_asignado': rule.exogena_format,
                'estado': 'procesado'
            })
        return processed, errors

    def process_file(self, file_path: str) -> Tuple[List[Dict], ProcessingResult]:
        """Procesa el archivo CSV."""
        df = self._load_csv(file_path)
        column_map = self.column_mapping if self.column_mapping else self._detect_column_mapping(df)
        df_norm = self._normalize_dataframe(df, column_map)

        rules = self.db.query(MappingRule).all()
        rule_map = {rule.puc_code: rule for rule in rules}

        processed_data, errors = self._process_rows(df_norm, rule_map)

        result = ProcessingResult(
            filename=Path(file_path).name,
            total_rows=len(df_norm),
            processed_rows=len(processed_data),
            error_rows=len(errors),
            errors=errors,
            timestamp=datetime.now()
        )
        return processed_data, result

    def generate_sample_csv(self) -> str:
        """Genera un CSV de ejemplo."""
        data = {
            'nit_tercero': ['900123456-1', '900234567-2', '900345678-3'],
            'nombre_tercero': ['Empresa A', 'Empresa B', 'Empresa C'],
            'valor': [1000000, 2500000, 500000],
            'codigo_puc': ['512010', '511005', '413505']
        }
        df = pd.DataFrame(data)
        file_path = Path('uploads') / 'sample_data.csv'
        file_path.parent.mkdir(exist_ok=True)
        df.to_csv(file_path, index=False)
        return str(file_path)

    def get_column_suggestions(self, file_path: str) -> Dict:
        """Analiza y sugiere columnas."""
        df = self._load_csv(file_path)
        try:
            detection = self._detect_column_mapping(df)
        except ValueError as e:
            detection = {'error': str(e)}
        return {
            'columns_available': df.columns.tolist(),
            'detected_mapping': detection,
            'sample_rows': df.head(3).to_dict('records')
        }

    def generate_excel(self, processed_data: List[Dict], file_id: str) -> str:
        """Genera un archivo Excel con detalle y resumen por formato."""
    
        
        # Crear workbook
        wb = Workbook()
        
        # --- Hoja 1: Detalle ---
        ws_detalle = wb.active
        ws_detalle.title = "Detalle"
        
        # Convertir a DataFrame para facilidad
        df_detalle = pd.DataFrame(processed_data)
        
        # Columnas amigables para el contador
        column_mapping = {
            'nit_tercero': 'NIT',
            'nombre_tercero': 'Nombre Tercero',
            'valor': 'Valor',
            'codigo_puc': 'Código PUC',
            'concepto_asignado': 'Concepto DIAN',
            'formato_asignado': 'Formato',
            'estado': 'Estado'
        }
        df_detalle = df_detalle.rename(columns=column_mapping)
        
        # Escribir encabezados
        for col_idx, col_name in enumerate(df_detalle.columns, 1):
            cell = ws_detalle.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Escribir datos
        for r_idx, row in enumerate(df_detalle.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws_detalle.cell(row=r_idx, column=c_idx, value=value)
        
        # Ajustar ancho de columnas
        for col in ws_detalle.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detalle.column_dimensions[column].width = adjusted_width
        
        # --- Hoja 2: Resumen por Formato ---
        ws_resumen = wb.create_sheet("Resumen por Formato")
        
        # Agrupar por formato y concepto
        df_resumen = df_detalle.groupby(['Formato', 'Concepto DIAN'])['Valor'].sum().reset_index()
        df_resumen['Cantidad'] = df_detalle.groupby(['Formato', 'Concepto DIAN']).size().values
        
        # Ordenar por formato
        df_resumen = df_resumen.sort_values('Formato')
        
        # Escribir encabezados
        for col_idx, col_name in enumerate(df_resumen.columns, 1):
            cell = ws_resumen.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Escribir datos
        for r_idx, row in enumerate(df_resumen.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws_resumen.cell(row=r_idx, column=c_idx, value=value)
        
        # Ajustar ancho
        for col in ws_resumen.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_resumen.column_dimensions[column].width = adjusted_width
        
        # --- Guardar archivo ---
        excel_dir = Path("outputs/excel")
        excel_dir.mkdir(parents=True, exist_ok=True)
        file_path = excel_dir / f"{file_id}_resumen.xlsx"
        wb.save(file_path)
        
        return str(file_path)
    
    def generate_xml(self, processed_data: List[Dict], file_id: str, format_code: str = "1001") -> str:
        """
        Genera un archivo XML para la DIAN según el formato especificado.
        Por ahora solo implementa el formato 1001 (pagos y retenciones).
        """
        from lxml import etree
        
        # Filtramos los registros que corresponden al formato solicitado
        filtered_data = [d for d in processed_data if d.get('formato_asignado') == format_code]
        
        if not filtered_data:
            raise ValueError(f"No hay datos para el formato {format_code}")
        
        # Crear el elemento raíz según el formato
        # Para 1001: InformacionExogena
        root = etree.Element("InformacionExogena")
        etree.SubElement(root, "Version").text = "1.0"
        etree.SubElement(root, "Año").text = "2025"  # Podríamos extraer del CSV
        etree.SubElement(root, "Periodo").text = "01"  # Por defecto
        
        # Crear el nodo de detalle
        detalle = etree.SubElement(root, "Detalle")
        
        for item in filtered_data:
            registro = etree.SubElement(detalle, "Registro")
            
            # Campos comunes para 1001
            nit = etree.SubElement(registro, "NitTercero")
            nit.text = item.get('nit_tercero', '')
            
            nombre = etree.SubElement(registro, "NombreTercero")
            nombre.text = item.get('nombre_tercero', '')
            
            valor = etree.SubElement(registro, "Valor")
            valor.text = str(item.get('valor', 0))
            
            # Aquí irían más campos según el concepto (retenciones, etc.)
            # Por simplicidad, solo agregamos lo básico.
            # En la práctica, necesitarías mapear cada concepto a su estructura específica.
        
        # Guardar XML
        xml_dir = Path("outputs/xml")
        xml_dir.mkdir(parents=True, exist_ok=True)
        file_path = xml_dir / f"{file_id}_{format_code}.xml"
        
        # Escribir el árbol XML con declaración y pretty print
        xml_str = etree.tostring(root, pretty_print=True, xml_declaration=True, encoding='UTF-8')
        with open(file_path, 'wb') as f:
            f.write(xml_str)
        
        return str(file_path)